from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox, ttk

def get_student_scores(file_name="students_scores.xlsx"):
    """
    Reads student scores from the Excel spreadsheet and returns them as a list of dictionaries.

    Args:
        file_name (str): The name of the Excel file containing student scores.

    Returns:
        list: A list of dictionaries, where each dictionary represents a student's record.
    """
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active

        # Extract the header row
        headers = [cell.value for cell in sheet[1]]

        # Extract the data rows
        student_scores = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            student_scores.append(dict(zip(headers, row)))

        return student_scores
    except FileNotFoundError:
        messagebox.showerror("Error", f"The file '{file_name}' does not exist.")
        return []
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")
        return []

def display_scores_ui():
    """
    Creates a UI to display student scores from the Excel spreadsheet.
    """
    def load_scores():
        scores = get_student_scores()
        if scores:
            # Clear the treeview
            for row in tree.get_children():
                tree.delete(row)

            # Insert data into the treeview
            for record in scores:
                tree.insert("", "end", values=(
                    record.get("Date", ""),
                    record.get("Student Name", ""),
                    record.get("Game Type", ""),
                    record.get("Correct Answers", ""),
                    record.get("Wrong Answers", ""),
                    record.get("Questions Asked", "")
                ))
        else:
            messagebox.showinfo("Info", "No scores found or unable to read the file.")

    def on_row_double_click(event):
        # Get the selected row
        selected_item = tree.selection()
        if selected_item:
            record = tree.item(selected_item, "values")
            if record:
                # Extract details from the record
                date_time = record[0]
                student_name = record[1]
                game_type = record[2]
                correct_answers = record[3]
                wrong_answers = record[4]
                questions_asked = record[5]

                # Split date and time
                date, time = date_time.split(" ")

                # Open a new window to display details
                detail_window = tk.Toplevel(root)
                detail_window.title("Student Details")

                tk.Label(detail_window, text=f"Student: {student_name}", font=("Arial", 12)).pack(pady=5)
                tk.Label(detail_window, text=f"Date: {date}", font=("Arial", 12)).pack(pady=5)
                tk.Label(detail_window, text=f"Time: {time}", font=("Arial", 12)).pack(pady=5)
                tk.Label(detail_window, text=f"Game: {game_type}", font=("Arial", 12)).pack(pady=5)
                tk.Label(detail_window, text=f"Correct Answers: {correct_answers}", font=("Arial", 12)).pack(pady=5)
                tk.Label(detail_window, text=f"Wrong Answers: {wrong_answers}", font=("Arial", 12)).pack(pady=5)
                tk.Label(detail_window, text="Questions Asked:", font=("Arial", 12, "bold")).pack(pady=5)

                # Display each question in a separate label
                for question in questions_asked.split("; "):
                    tk.Label(detail_window, text=f"- {question}", font=("Arial", 10)).pack(anchor="w", padx=10)

    # Create the main window
    root = tk.Tk()
    root.title("Student Scores")

    # Create a frame for the treeview
    frame = tk.Frame(root)
    frame.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

    # Create the treeview
    columns = ("Date", "Student Name", "Game Type", "Correct Answers", "Wrong Answers", "Questions Asked")
    tree = ttk.Treeview(frame, columns=columns, show="headings")
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    # Define column headings
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, width=150, anchor="center")

    # Add a scrollbar
    scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    # Add a button to load scores
    load_button = tk.Button(root, text="Load Scores", command=load_scores)
    load_button.pack(pady=10)

    # Bind double-click event to the treeview
    tree.bind("<Double-1>", on_row_double_click)

    # Run the main loop
    root.mainloop()

# Run the UI
if __name__ == "__main__":
    display_scores_ui()