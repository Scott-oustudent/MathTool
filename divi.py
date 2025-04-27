import random
from openpyxl import Workbook, load_workbook
from datetime import datetime

def save_to_excel(student_name, game_type, correct_ans, wrong_ans, questions):
    file_name = "students_scores.xlsx"
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Student Name", "Game Type", "Correct Answers", "Wrong Answers", "Questions Asked"])

    questions_str = "; ".join(questions)
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([date, student_name, game_type, correct_ans, wrong_ans, questions_str])
    workbook.save(file_name)

def division_game(student_name):
    correct_ans = 0
    wrong_ans = 0
    questions = []

    print("\nWelcome to the Division Quiz!")
    counter = 0
    while counter < 10:
        num1 = random.randint(2, 12)
        num2 = random.randint(1, 10)
        answer = round(num1 / num2, 2)
        question = f"{num1} ÷ {num2}"

        try:
            stu_ans = float(input(f"What is {question} (round to 2 decimal places): "))
            if stu_ans == answer:
                correct_ans += 1
                print("Correct. Well done!!!")
            else:
                wrong_ans += 1
                print(f"No sorry. The correct answer was {answer}")
            questions.append(f"{question} = {stu_ans} (Correct: {stu_ans == answer})")
            counter += 1
        except ValueError:
            print("Please enter a valid number.")
            questions.append(f"{question} = Invalid Input")

    print("\nDivision Quiz Summary:")
    print(f"Correct Answers: {correct_ans}")
    print(f"Wrong Answers: {wrong_ans}")
    save_to_excel(student_name, "Division", correct_ans, wrong_ans, questions)
