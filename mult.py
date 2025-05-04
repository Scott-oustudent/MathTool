import random
from openpyxl import Workbook, load_workbook
from datetime import datetime
import tkinter as tk
from tkinter import messagebox

def save_to_excel(student_name, game_type, correct_ans, wrong_ans, questions):
    file_name = "students_scores.xlsx"
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Student Name", "Game Type", "Correct Answers", "Wrong Answers", "Questions Asked"])

    questions_str = "; ".join(questions)
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([date, student_name, game_type, correct_ans, wrong_ans, questions_str])
    workbook.save(file_name)

def mult_game_ui(student_name):
    def start_game():
        nonlocal work
        try:
            work = int(table_entry.get())
            if work < 1 or work > 10:
                raise ValueError
            table_entry_frame.pack_forget()
            next_question()
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter a number between 1 and 10.")

    def next_question():
        nonlocal counter, correct_ans, wrong_ans
        if counter >= 10:
            messagebox.showinfo("Quiz Complete", f"Correct Answers: {correct_ans}\nWrong Answers: {wrong_ans}")
            save_to_excel(student_name, "Multiplication", correct_ans, wrong_ans, questions)
            root.destroy()
            return

        num = random.randint(1, 10)
        question = f"{work} x {num}"
        while question in asked_questions:
            num = random.randint(1, 10)
            question = f"{work} x {num}"
        asked_questions.add(question)

        answer = work * num
        questions.append(question)

        def check_answer():
            nonlocal correct_ans, wrong_ans, counter
            try:
                stu_ans = int(answer_entry.get())
                if stu_ans == answer:
                    correct_ans += 1
                    result_label.config(text="Correct! Well done!", fg="green")
                else:
                    wrong_ans += 1
                    result_label.config(text=f"Wrong! The correct answer was {answer}.", fg="red")
                questions[counter] += f" = {stu_ans} (Correct: {stu_ans == answer})"
                counter += 1
                next_question()
            except ValueError:
                result_label.config(text="Please enter a valid number.", fg="red")

        question_label.config(text=f"What is {question}?")
        answer_entry.delete(0, tk.END)
        submit_button.config(command=check_answer)

    root = tk.Tk()
    root.title("Multiplication Quiz")

    correct_ans = 0
    wrong_ans = 0
    counter = 0
    questions = []
    asked_questions = set()
    work = 0

    tk.Label(root, text=f"Welcome {student_name} to the Multiplication Quiz!", font=("Arial", 14)).pack(pady=10)

    # Frame for selecting the multiplication table
    table_entry_frame = tk.Frame(root)
    table_entry_frame.pack(pady=10)
    tk.Label(table_entry_frame, text="Enter the multiplication table (1-10):").pack(side=tk.LEFT)
    table_entry = tk.Entry(table_entry_frame)
    table_entry.pack(side=tk.LEFT)
    tk.Button(table_entry_frame, text="Start", command=start_game).pack(side=tk.LEFT)

    # Frame for the quiz
    question_label = tk.Label(root, text="", font=("Arial", 12))
    question_label.pack(pady=10)

    answer_entry = tk.Entry(root)
    answer_entry.pack(pady=5)

    submit_button = tk.Button(root, text="Submit", command=None)
    submit_button.pack(pady=5)

    result_label = tk.Label(root, text="", font=("Arial", 10))
    result_label.pack(pady=10)

    root.mainloop()
