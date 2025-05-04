import random
from openpyxl import Workbook, load_workbook
from datetime import datetime
import tkinter as tk
from tkinter import messagebox

def save_to_excel(student_name, game_type, correct_ans, wrong_ans, questions):
    file_name = "students_scores.xlsx"
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Student Name", "Game Type", "Correct Answers", "Wrong Answers", "Questions Asked"])

    questions_str = "; ".join(questions)
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([date, student_name, game_type, correct_ans, wrong_ans, questions_str])
    workbook.save(file_name)

def subtraction_game_ui(student_name):
    def next_question():
        nonlocal counter, correct_ans, wrong_ans
        if counter >= 10:
            messagebox.showinfo("Quiz Complete", f"Correct Answers: {correct_ans}\nWrong Answers: {wrong_ans}")
            save_to_excel(student_name, "Subtraction", correct_ans, wrong_ans, questions)
            root.destroy()
            return

        num1 = random.randint(1, 50)
        num2 = random.randint(1, 50)
        if num1 < num2:
            num1, num2 = num2, num1  # Ensure no negative results
        answer = num1 - num2
        question = f"{num1} - {num2}"
        questions.append(question)

        def check_answer():
            nonlocal correct_ans, wrong_ans, counter
            try:
                stu_ans = int(answer_entry.get())
                if stu_ans == answer:
                    correct_ans += 1
                    result_label.config(text="Correct! Well done!", fg="green")
                else:
                    wrong_ans += 1
                    result_label.config(text=f"Wrong! The correct answer was {answer}.", fg="red")
                questions[counter] += f" = {stu_ans} (Correct: {stu_ans == answer})"
                counter += 1
                next_question()
            except ValueError:
                result_label.config(text="Please enter a valid number.", fg="red")

        question_label.config(text=f"What is {question}?")
        answer_entry.delete(0, tk.END)
        submit_button.config(command=check_answer)

    root = tk.Tk()
    root.title("Subtraction Quiz")

    correct_ans = 0
    wrong_ans = 0
    counter = 0
    questions = []

    tk.Label(root, text=f"Welcome {student_name} to the Subtraction Quiz!", font=("Arial", 14)).pack(pady=10)
    question_label = tk.Label(root, text="", font=("Arial", 12))
    question_label.pack(pady=10)

    answer_entry = tk.Entry(root)
    answer_entry.pack(pady=5)

    submit_button = tk.Button(root, text="Submit", command=None)
    submit_button.pack(pady=5)

    result_label = tk.Label(root, text="", font=("Arial", 10))
    result_label.pack(pady=10)

    next_question()
    root.mainloop()

def main():
    print("Welcome to the Math Quiz Program!")
    student_name = input("Enter your name: ")

    print("Choose a game:")
    print("1. Multiplication")
    print("2. Division")
    print("3. Subtraction")
    print("4. Addition")

    choice = input("Enter the number of your choice: ")

    if choice == "1":
        mult_game(student_name)
    elif choice == "2":
        division_game(student_name)
    elif choice == "3":
        subtraction_game_ui(student_name)
    elif choice == "4":
        addition_game(student_name)
    else:
        print("Invalid choice. Please restart the program.")

if __name__ == "__main__":
    main()
